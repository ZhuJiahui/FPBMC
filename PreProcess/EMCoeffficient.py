# -*- coding: utf-8 -*-

'''
Created on 2014-3-24

@author: ZhuJiahui506
'''


from __future__ import division
import os
import math 
from openpyxl import Workbook
from numpy import random
import numpy as np
import time
from datetime import datetime
from ExcelToolkit import open_sheet
from TextToolkit import quick_write_list_to_text

'''
step 2
Generate EM coefficients for the features
'''


pi = math.pi
exp = math.exp
pow = math.pow
sqrt = math.sqrt 
log = math.log

def get_data_list(read_filename):
    weibo_sheet = open_sheet(read_filename)
    weibo_row = weibo_sheet.nrows
    
    print 'Number of the Weibo row: %d' % weibo_row
    
    forwards = []
    comments = []
    urls = []
    
    for i in range(1, weibo_row):
        forwards.append(float(weibo_sheet.cell(i, 15).value))
        comments.append(float(weibo_sheet.cell(i, 16).value))
        urls.append(float(weibo_sheet.cell(i, 17).value))
    
    return [forwards, comments, urls]

def gaussD(x, mu, sigma):
    if sigma < 5:
#        if flag == 0:
#        print 'error'
#            flag = 1
        sigma = 3 + 2 * random.random()
    return exp(-1 * (x - mu) * (x - mu) / 2 / sigma) / sqrt(2 * pi * sigma)

def eStep(dataList, aList, muList, sigList , enlar=1e1):
    E_MAX = []
    rows = len(dataList[0])
    cols = len(dataList)
    gaussMatrix = []
    rowSumList = []
    Ln = 1
    
    for j in range(cols):
        gaussMatrix_j = []
        for i in range(rows):
            #放大
            tmp = aList[j] * gaussD(dataList[j][i], muList[j], sigList[j]) * enlar
            gaussMatrix_j.append(tmp)
        gaussMatrix.append(gaussMatrix_j)
       
    for i in range(rows):
        rowSum = 0
        for j in range(cols):
            rowSum += gaussMatrix[j][i]
        Ln *= rowSum
        if rowSum == 0:
            print i, j    
        rowSumList.append(rowSum)
        
    # 计算期望
    for j in range(cols):
        E_MAX_j = []
        for i in range(rows):
            tmpSum = rowSumList[i]
            
            if tmpSum == 0:
                wij = 1 / cols
            else:
                wij = gaussMatrix[j][i] / tmpSum
            E_MAX_j.append(wij)
        E_MAX.append(E_MAX_j)
    
#    for j in range(cols):
#        print E_MAX[j][21]
    return Ln, E_MAX

def mStep(E_MAX, dataList, aList, muList, sigList):
    rows = len(dataList[0])
    cols = len(dataList)
    
#    print aList, muList, sigList
    
    nAList = []
    nMuList = []
    nSigList = []
    # 最大化参数
    for j in range(cols):
        sum_a = 0
        sum_mu = 0
        sum_sig = 0
        for i in range(rows):
            sum_a += E_MAX[j][i]
            sum_mu += E_MAX[j][i] * dataList[j][i]
            sum_sig += E_MAX[j][i] * (dataList[j][i] - muList[j]) * (dataList[j][i] - muList[j])
        nA = sum_a / rows
        nAList.append(nA)
        if (aList[j] == 0):
            nMu = 0
            nSig = 1
        else:
            nMu = sum_mu / (rows * aList[j])
            nSig = sum_sig / (rows * aList[j])
        nMuList.append(nMu)
        
        nSigList.append(nSig)
    
    return nAList, nMuList, nSigList
        
def getMuAndSigma(list):
    mu = sum(list) / len(list)
    sigma = 0
    for i in list:
        sigma += (i - mu) * (i - mu)
    sigma = sigma / len(list)
    return mu, sigma

def getMSForRequired(datalist):
    muList = []
    sigList = []
    le = len(datalist)
    for i in range(le):
#        print i
        mu, sigma = getMuAndSigma(datalist[i])
        muList.append(mu)
        sigList.append(sigma)
    return muList, sigList

def cope(data, wf, row=0, le=5, threshold=1e-8, max_iter=100):
    #data = get_data_list(rf)
    le = len(data)
    a = []
    for i in range(le):
        a.append(1 / le)
    muList, sigList = getMSForRequired(data)
    count = 0
    oLn = 0
    muMatrix = [muList]
    sigMatrix = [sigList]
    aMatrix = [a]
    starttime = datetime.now()
    stepStarttime = datetime.now()
    stepEndTime = datetime.now()
    stepTimeTakes = 0
    stepTimeAvg = 0
    while count < max_iter:
        stepStarttime = datetime.now()
        count += 1
        Ln, E_Max = eStep(data, a , muList, sigList)
        newA, newMuList, newSigList = mStep(E_Max, data, a, muList, sigList)
        
#        if abs(Ln - oLn) <= threshold:
#            break
        
        aCha = 0
        muCha = 0
        sigCha = 0
        for i in range(le):
            aCha += (newA[i] - a[i]) * (newA[i] - a[i])
            muCha += (newMuList[i] - muList[i]) * (newMuList[i] - muList[i])
            sigCha += (newSigList[i] - muList[i]) * (newSigList[i] - muList[i])
            
        if aCha / le < threshold or muCha / le < threshold or sigCha / le < threshold:
            break
            
        oLn = Ln
        a = newA
        muList = newMuList
        sigList = newSigList
        muMatrix.append(muList)
        sigMatrix.append(sigList)
        aMatrix.append(a)
        stepTimeTakes = (datetime.now() - stepStarttime)
#        print oLn
    
    timeTakes = (datetime.now() - starttime)
    stepTimeAvg = (timeTakes.seconds * 1e7 + timeTakes.microseconds) / count / 1e7
    print '迭代次数为 : %d' % (count)
    print '本次EM处理时间为 ： %d 秒' % (timeTakes.seconds) 
    print '本次EM每次迭代处理时间为 ： %f 秒' % (stepTimeAvg)   
    
    sorted = getTopN(a, data)
    firstRow = list(a)
    firstRow.append(u'每步处理时间')
    firstRow.append(stepTimeAvg)
    writeToExcel(sorted, firstRow , wf);
    return list(a)


def getTopN(a, data, N=50, enlar=1):
    new_data = []
    le = len(data)
    for i in range(len(data[0])):
        tmpSum = 0
        tmpList = [i + 2]
        for j in range(le):
            tmpSum += data[j][i] * a[j]
            tmpList.append(data[j][i])
        tmpList.insert(1, tmpSum)
        tmp = tuple(tmpList)
        new_data.append(tmp)
#    result = sorted(new_data, key=lambda x:x[1], reverse=True)
#    for i in range(N):
#        print result[i]
    return new_data

def getNData(a, data):
    new_data = []
    le = len(data)
    for i in range(len(data[0])):
        tmpList = []
        sum = 0
        for j in range(le):
            tmpList.append(data[j][i])
            sum += data[j][i] * a[j]
#        tmpList.insert(le+1, sum)
        tmp = tuple(tmpList) 
        new_data.append(tmp)
#    result = sorted(new_data, key=lambda x:x[1], reverse=True)
    result = new_data
    return result

def writeToExcel(data, a, f):
    wb = Workbook(optimized_write=True);
    ws = wb.create_sheet(0, 'sheet1')
    ws.append(a)
    ws.append(['ExcelID', 'EM值'])
    rows = len(data)
    cols = len(data[0])
    for i in range(rows):
        writeRow = []
        for j in range(cols):
            writeRow.append(data[i][j])
        ws.append(writeRow)
    wb.save(f)
        
def main(): 
    print 

if __name__ == '__main__':
     
    start = time.clock()
    now_directory = os.getcwd()
    root_directory = os.path.dirname(now_directory) + '/'
    
    read_filename = root_directory + u'dataset/em/score.txt'
    write_directory = root_directory + u'dataset/em10'
    write_filename = write_directory + u'/em_coefficients.txt'
    
    #threshold = 1e-5
    #max_iter = 100
    iter_number = 10
    
    if (not(os.path.exists(write_directory))):
        os.mkdir(write_directory)

        
    forwards = []
    comments = []
    urls = []
        
    f = open(read_filename, 'r')
    line = f.readline()
    while line:
        each_line = line.split()
        forwards.append(float(each_line[0]))
        comments.append(float(each_line[1]))
        urls.append(float(each_line[2]))
            
        line = f.readline()
    f.close()
        
    forwards_coefficients = []
    comments_coefficients = []
    urls_coefficients = []
        
    for j in range(iter_number):
        # 由于数据量大，产生一个随机样本
        sample_number = 10000
        random_series = np.random.random_integers(0, len(forwards) - 1, sample_number)
            
        em_forwards = []
        em_comments = []
        em_urls = []
            
        for each in random_series:
            em_forwards.append(forwards[each])
            em_comments.append(comments[each])
            em_urls.append(urls[each])
            
        coefficients = cope([em_forwards, em_comments, em_urls], write_directory + '/em_process' + str(j + 1) + '.xlsx', row=0, le=5, threshold=1e-8, max_iter=100)

        forwards_coefficients.append(coefficients[0])
        comments_coefficients.append(coefficients[1])
        urls_coefficients.append(coefficients[2])
        
    average_coefficients = [np.average(forwards_coefficients), np.average(comments_coefficients), np.average(urls_coefficients)]
    print average_coefficients
        
    average_coefficients_to_string = [str(x) for x in average_coefficients]
    quick_write_list_to_text(average_coefficients_to_string, write_filename)
    
    print 'Total time %f seconds' % (time.clock() - start)
    print 'Complete !!!'
    